"""
XLSX and XLSM parser using openpyxl.

Emits one section per data row rather than one blob per sheet.

A sheet flattened into a single text block gets split on character count, so a
record can be cut in half and every chunk after the first loses the header row
entirely — leaving rows of bare values with nothing saying which column each
belongs to. Pairing every value with its column name, one record at a time, is
what makes "expense ratio for Fund X" match Fund X's row instead of a slab of
run-together cells.
"""

import structlog
from openpyxl import load_workbook

from app.ingestion.parser import DocumentParser, ParsedDocument, ParsedSection

logger = structlog.get_logger()

# Rows are scanned for a header within this many leading rows; sheets often
# open with a title or a blank line before the real column names.
HEADER_SEARCH_DEPTH = 8
# Guards against pathologically wide sheets producing unreadable records.
MAX_COLUMNS = 40
# Below this, a sheet is treated as prose rather than a table.
MIN_TABLE_ROWS = 2
# One chunk per row is ideal for retrieval but only while rows are countable.
# Workbooks routinely carry a raw dump sheet — this one has 15,000 rows in
# "Sheet1" — and a chunk each would swamp the index and dominate every search.
# Past this many rows a sheet is packed into blocks instead, which keeps the
# data searchable with a bounded number of chunks.
ROW_LEVEL_MAX_ROWS = 300
# Character budget per packed block, matched to the chunker's ceiling.
BLOCK_CHAR_BUDGET = 1000
# A hard stop so a pathological sheet cannot stall ingestion outright.
MAX_ROWS_PER_SHEET = 6000


def _clean(value) -> str:
    return "" if value is None else str(value).strip()


def _pick_header(rows: list[list[str]]) -> int:
    """
    Index of the most plausible header row.

    The header is taken to be the row with the most distinct non-empty cells
    among the first few — which beats "row 0" on the many real sheets that
    begin with a title, a date stamp, or a blank spacer.
    """
    best_index, best_score = 0, -1
    for i, row in enumerate(rows[:HEADER_SEARCH_DEPTH]):
        filled = [c for c in row if c]
        score = len(set(filled))
        # A header is text, not numbers.
        if filled and all(not c.replace(".", "", 1).replace("-", "", 1).isdigit() for c in filled):
            score += 2
        if score > best_score:
            best_index, best_score = i, score
    return best_index


class XLSXParser(DocumentParser):

    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xlsm", ".xltx", ".xltm"]

    async def parse(self, file_path: str, filename: str) -> ParsedDocument:
        wb = load_workbook(file_path, data_only=True, read_only=True, keep_vba=False)
        sections: list[ParsedSection] = []
        tables: list[dict] = []
        sheet_names = list(wb.sheetnames)

        try:
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                rows_data: list[list[str]] = []

                for row in ws.iter_rows(values_only=True):
                    values = [_clean(cell) for cell in row][:MAX_COLUMNS]
                    if any(values):
                        rows_data.append(values)

                if not rows_data:
                    continue

                if len(rows_data) < MIN_TABLE_ROWS:
                    # Not enough structure to treat as a table.
                    sections.append(ParsedSection(
                        content=f"Sheet: {sheet_name}\n" + "\n".join(
                            " | ".join(r) for r in rows_data
                        ),
                        section=sheet_name,
                    ))
                    continue

                header_index = _pick_header(rows_data)
                headers = rows_data[header_index]
                # Unlabelled columns still need a stable name.
                headers = [h or f"Column {i + 1}" for i, h in enumerate(headers)]
                body = rows_data[header_index + 1:]

                if len(body) > MAX_ROWS_PER_SHEET:
                    logger.warning(
                        "sheet_truncated",
                        sheet=sheet_name,
                        rows=len(body),
                        kept=MAX_ROWS_PER_SHEET,
                    )
                    body = body[:MAX_ROWS_PER_SHEET]

                records: list[tuple[str, str, int]] = []
                for offset, row in enumerate(body):
                    pairs = [
                        f"{headers[i]}: {value}"
                        for i, value in enumerate(row)
                        if i < len(headers) and value
                    ]
                    if not pairs:
                        continue
                    # The first populated cell is almost always the record's
                    # identity (a fund name, a ticker), so it leads the text and
                    # also keys the section.
                    label = next((v for v in row if v), f"Row {offset + 1}")
                    records.append((label, "\n".join(pairs), header_index + offset + 2))

                if len(records) <= ROW_LEVEL_MAX_ROWS:
                    # One chunk per record — the precise case. A distinct
                    # section per record also stops the chunker merging
                    # neighbouring records back together.
                    for label, body_text, row_number in records:
                        sections.append(ParsedSection(
                            content=f"{sheet_name} — {label}\n{body_text}",
                            section=f"{sheet_name} · {label}"[:500],
                            metadata={
                                "record_label": label,
                                "sheet": sheet_name,
                                "row_number": row_number,
                            },
                        ))
                else:
                    # Pack records into blocks. Each block still carries its
                    # column names, so values never lose their labels.
                    block: list[str] = []
                    block_chars = 0
                    block_index = 0
                    first_row = records[0][2] if records else 0

                    def emit(buf: list[str], start_row: int, index: int) -> None:
                        if not buf:
                            return
                        sections.append(ParsedSection(
                            content=f"{sheet_name} (rows from {start_row})\n\n"
                                    + "\n\n".join(buf),
                            section=f"{sheet_name} · block {index}"[:500],
                            metadata={
                                "sheet": sheet_name,
                                "row_number": start_row,
                                "packed_records": len(buf),
                            },
                        ))

                    for label, body_text, row_number in records:
                        record_text = f"{label}\n{body_text}"
                        if block and block_chars + len(record_text) > BLOCK_CHAR_BUDGET:
                            emit(block, first_row, block_index)
                            block_index += 1
                            block, block_chars = [], 0
                            first_row = row_number
                        block.append(record_text)
                        block_chars += len(record_text)
                    emit(block, first_row, block_index)

                tables.append({
                    "sheet": sheet_name,
                    "headers": headers,
                    "rows": body,
                })
        finally:
            wb.close()

        return ParsedDocument(
            sections=sections,
            title=filename,
            metadata={"format": filename.split(".")[-1].lower(), "sheets": len(sheet_names)},
            tables=tables,
        )
